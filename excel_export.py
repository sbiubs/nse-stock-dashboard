"""
Builds a formatted, color-coded multi-sheet Excel workbook from the
day's screening results — mirrors the visual style used across your
other GST/finance tools (color-coded status, clean headers).
"""
import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BUY_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
SELL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HOLD_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

SIGNAL_FILLS = {"BUY": BUY_FILL, "SELL": SELL_FILL, "HOLD": HOLD_FILL}


def _write_sheet(ws, df: pd.DataFrame, signal_col: str = None):
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    signal_col_idx = list(df.columns).index(signal_col) + 1 if signal_col in df.columns else None

    for row in df.itertuples(index=False):
        ws.append(list(row))
        if signal_col_idx:
            signal_val = getattr(row, signal_col) if hasattr(row, signal_col) else None
            fill = SIGNAL_FILLS.get(signal_val)
            if fill:
                ws.cell(row=ws.max_row, column=signal_col_idx).fill = fill

    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 45)


def build_workbook(full_results: pd.DataFrame, top_calls: dict) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Short-Term Buys"
    _write_sheet(ws1, top_calls["short_term_buys"], signal_col="short_term_signal")

    ws2 = wb.create_sheet("Short-Term Sells")
    _write_sheet(ws2, top_calls["short_term_sells"], signal_col="short_term_signal")

    ws3 = wb.create_sheet("Long-Term Buys")
    _write_sheet(ws3, top_calls["long_term_buys"], signal_col="long_term_signal")

    ws4 = wb.create_sheet("Long-Term Sells")
    _write_sheet(ws4, top_calls["long_term_sells"], signal_col="long_term_signal")

    ws5 = wb.create_sheet("Full Screen")
    _write_sheet(ws5, full_results, signal_col="short_term_signal")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
